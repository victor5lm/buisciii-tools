import os
import glob
import re
import shlex
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

# ------------------------------------------------------
# File Paths
# ------------------------------------------------------

# Paths to input files and directories
xlsx_template = "summary_outbreak_template.xlsx"
samples_file = "../ANALYSIS/samples_id.txt"
tsv_file = glob.glob("../ANALYSIS/*_CHARACTERIZATION/99-stats/ariba_mlst_full.tsv")
csv_file = glob.glob("../ANALYSIS/*_ASSEMBLY/99-stats/kmerfinder_summary.csv")
mapping_file = glob.glob("../ANALYSIS/*_SNIPPY/99-stats/mapping_stats_summary.txt")
wgs_metrics_file = glob.glob(
    "../ANALYSIS/*_SNIPPY/99-stats/wgs_metrics_all_filtered.txt"
)
variants_stats_file = glob.glob("../ANALYSIS/*_SNIPPY/99-stats/variants_stats.txt")
quast_report_file = glob.glob(
    "../ANALYSIS/*_ASSEMBLY/03-assembly/quast/global_report/transposed_report.tsv"
)
quast_dir = glob.glob("../ANALYSIS/*_ASSEMBLY/03-assembly/quast/per_reference_reports/")
virulence_file = glob.glob(
    "../ANALYSIS/*_CHARACTERIZATION/99-stats/ariba_vfdb_full.csv"
)
card_file = glob.glob("../ANALYSIS/*_CHARACTERIZATION/99-stats/ariba_card.csv")
amrfinder_dir = glob.glob("../ANALYSIS/*_CHARACTERIZATION/03-amrfinderplus")
plasmid_file = glob.glob(
    "../ANALYSIS/*_PLASMIDID/NO_GROUP/NO_GROUP_final_results_per_sample.tab"
)
mlva_file = glob.glob(
    "../ANALYSIS/*_CHARACTERIZATION/05-mlva/MLVA_output/MLVA_analysis_assemblies.csv"
)
phylo_alignment_file = glob.glob("../ANALYSIS/*_SNIPPY/05-snippy/phylo.aln")
nonrecombinant_alignment_file = glob.glob(
    "../ANALYSIS/*_SNIPPY/05-snippy/clean.core.aln"
)
snippy_commands_file = glob.glob("../ANALYSIS/*_SNIPPY/05-snippy/commands.out")


def get_first_match(file_list):
    return file_list[0] if file_list else None


def normalize_sample_id(sample_id):
    """Return the sample identifier used in ``samples_id.txt``.

    Analysis programs add pipeline-specific suffixes to the sample name.  Keeping
    the normalization in one place prevents otherwise valid results from being
    skipped when they are matched to workbook rows.
    """
    sample_id = os.path.basename(str(sample_id).strip())
    sample_id = re.sub(
        r"(?:-unicycler)?\.scaffolds(?:\.(?:fa|fasta|fna)(?:\.gz)?)?$",
        "",
        sample_id,
        flags=re.IGNORECASE,
    )
    sample_id = re.sub(r"_S$", "", sample_id, flags=re.IGNORECASE)
    return sample_id


tsv_file = get_first_match(tsv_file)
csv_file = get_first_match(csv_file)
mapping_file = get_first_match(mapping_file)
wgs_metrics_file = get_first_match(wgs_metrics_file)
variants_stats_file = get_first_match(variants_stats_file)
quast_report_file = get_first_match(quast_report_file)
quast_dir = get_first_match(quast_dir)
virulence_file = get_first_match(virulence_file)
card_file = get_first_match(card_file)
amrfinder_dir = get_first_match(amrfinder_dir)
plasmid_file = get_first_match(plasmid_file)
mlva_file = get_first_match(mlva_file)
phylo_alignment_file = get_first_match(phylo_alignment_file)
nonrecombinant_alignment_file = get_first_match(nonrecombinant_alignment_file)
snippy_commands_file = get_first_match(snippy_commands_file)

# ------------------------------------------------------
# File Reading Functions
# ------------------------------------------------------


def read_samples(file_path):
    """
    Reads sample identifiers from a file and maps them to row numbers for the Excel sheet.

    Args:
        file_path (str): Path to the file containing sample identifiers.

    Returns:
        dict: A dictionary where keys are sample identifiers and values are their corresponding row numbers.
    """
    samples = {}
    with open(file_path, "r") as f:
        for index, line in enumerate(f, start=3):  # Start at row 3 in Excel
            sample_id = normalize_sample_id(line)
            samples[sample_id] = index
    return samples


def read_ariba_mlst(tsv_file):
    """
    Processes the MLST TSV file and extracts relevant sequence typing data.

    Args:
        tsv_file (str): Path to the MLST results file.

    Returns:
        tuple: A dictionary mapping sample IDs to their sequence types (ST) and gene sequences,
               and a header string for the gene columns.
    """
    df_mlst = pd.read_csv(tsv_file, sep="\t")
    mlst_columns = df_mlst.columns[2:].tolist()
    mlst_header = "/".join(mlst_columns)
    mlst_dict = {
        str(row["sample_id"]): {
            "ST": row["ST"],
            "genes": "/".join(map(str, row[mlst_columns].values)),
        }
        for _, row in df_mlst.iterrows()
    }
    return mlst_dict, mlst_header


def read_kmerfinder(csv_file):
    """
    Reads a CSV file and extracts specific columns for each sample.

    Args:
        csv_file (str): Path to the CSV file.

    Returns:
        dict: A dictionary where keys are sample names and values are dictionaries containing
              selected columns ('colE', 'colF', 'colG').
    """
    try:
        df_kmerfinder = pd.read_csv(csv_file)

        # Ensure the expected columns exist
        expected_columns = [
            "sample_name",
            df_kmerfinder.columns[1],
            df_kmerfinder.columns[2],
            df_kmerfinder.columns[4],
        ]
        for col in expected_columns:
            if col not in df_kmerfinder.columns:
                raise KeyError(f"Missing expected column: {col}")

        return {
            normalize_sample_id(row["sample_name"]): {
                "colE": row.iloc[1],
                "colF": row.iloc[2],
                "colG": row.iloc[4],
            }
            for _, row in df_kmerfinder.iterrows()
        }

    except FileNotFoundError:
        print(f"Error: The file '{csv_file}' was not found.")
        return {}
    except pd.errors.EmptyDataError:
        print(f"Error: The file '{csv_file}' is empty.")
        return {}
    except pd.errors.ParserError:
        print(f"Error: The file '{csv_file}' could not be parsed.")
        return {}
    except KeyError as e:
        print(f"Error: {e}")
        return {}


def read_mapping_stats(mapping_file):
    """
    Reads a mapping statistics file and extracts mapping values per sample.

    Args:
        mapping_file (str): Path to the mapping statistics file.

    Returns:
        dict: A dictionary where keys are sample names and values are mapping statistics.
    """
    try:
        df_mapping = pd.read_csv(mapping_file, sep=";")
        df_mapping.columns = df_mapping.columns.str.strip()

        if (
            "SAMPLENAME" not in df_mapping.columns
            or "MAPPING" not in df_mapping.columns
        ):
            raise KeyError("Missing required columns: 'SAMPLENAME' or 'MAPPING'")

        return {
            str(row["SAMPLENAME"]): row["MAPPING"] for _, row in df_mapping.iterrows()
        }
    except FileNotFoundError:
        print(f"Error: The file '{mapping_file}' was not found.")
        return {}
    except pd.errors.EmptyDataError:
        print(f"Error: The file '{mapping_file}' is empty.")
        return {}
    except pd.errors.ParserError:
        print(f"Error: The file '{mapping_file}' could not be parsed.")
        return {}
    except KeyError as e:
        print(f"Error: {e}")
        return {}


def read_wgs_metrics(wgs_metrics_file):
    """
    Reads a Whole Genome Sequencing (WGS) metrics file and extracts relevant coverage metrics.

    Args:
        wgs_metrics_file (str): Path to the WGS metrics file.

    Returns:
        dict: A dictionary where keys are sample names and values are dictionaries containing
              mean coverage and percentage of reads above 10X coverage.
    """
    try:
        df_wgs = pd.read_csv(wgs_metrics_file, sep="\t", dtype={"SAMPLENAME": str})

        if (
            "SAMPLENAME" not in df_wgs.columns
            or "MEAN_COVERAGE" not in df_wgs.columns
            or "PCT_10X" not in df_wgs.columns
        ):
            raise KeyError(
                "Missing required columns: 'SAMPLENAME', 'MEAN_COVERAGE', or 'PCT_10X'"
            )

        return {
            str(row["SAMPLENAME"]): {
                "MEAN_COVERAGE": row["MEAN_COVERAGE"],
                "PCT_10X": row["PCT_10X"],
            }
            for _, row in df_wgs.iterrows()
        }
    except Exception as e:
        print(f"Error processing WGS metrics file: {e}")
        return {}


def read_virulence_stats(virulence_file):
    """
    Reads a virulence gene statistics file and processes the data.

    Args:
        virulence_file (str): Path to the virulence statistics file.

    Returns:
        dict: A dictionary where keys are sample names and values are dictionaries containing:
              - "genes": A string with detected virulence genes.
              - "count": The number of virulence genes detected.
    """
    df_virulence = pd.read_csv(virulence_file)

    virulence_dict = {}
    for _, row in df_virulence.iterrows():
        sample = str(row["sample"])

        # Process the virulence gene list, removing unnecessary characters
        virulence_genes = row["virulence"].strip("[]").replace("'", "").split(", ")
        virulence_genes_cleaned = [
            gene.replace(".match", "") for gene in virulence_genes
        ]
        virulence_list = ", ".join(virulence_genes_cleaned)

        # Get the virulence gene count
        virulence_count = row["virulence_genes_vfdb"]

        virulence_dict[sample] = {"genes": virulence_list, "count": virulence_count}

    return virulence_dict


def read_card_stats(card_file):
    """
    Reads a CARD (Comprehensive Antibiotic Resistance Database) file and extracts resistance gene data.

    Args:
        card_file (str): Path to the CARD resistance gene statistics file.

    Returns:
        dict: A dictionary where keys are sample names and values are dictionaries containing:
              - "genes": A string with detected resistance genes.
    """
    df_card = pd.read_csv(card_file)

    card_dict = {}
    for _, row in df_card.iterrows():
        sample = str(row["sample"])

        # Process the resistance gene list, removing unnecessary characters
        card_genes = (
            row["resistance_genes_card"].strip("[]").replace("'", "").split(", ")
        )
        card_genes_cleaned = [gene.replace(".match", "") for gene in card_genes]
        card_list = ", ".join(card_genes_cleaned)

        card_dict[sample] = {"genes": card_list}

    return card_dict


def read_amrfinder_results(directory):
    """
    Reads AMRFinder results from multiple TSV files in a given directory.

    Args:
        directory (str): Path to the directory containing AMRFinder result files.

    Returns:
        dict: A dictionary where keys are sample names and values are strings of detected resistance genes.
    """
    resistance_dict = {}

    for filename in sorted(os.listdir(directory)):
        if filename.endswith("_out.tsv"):
            sample_id = normalize_sample_id(filename.replace("_out.tsv", ""))
            file_path = os.path.join(directory, filename)

            df = pd.read_csv(file_path, sep="\t", keep_default_na=False)

            # AMRFinderPlus renamed these fields in recent releases.
            gene_column = next(
                (column for column in ("Element symbol", "Gene symbol") if column in df.columns),
                None,
            )
            if gene_column:
                # The output can also contain virulence/stress hits; only AMR
                # elements belong in the resistance summary.
                type_column = next(
                    (column for column in ("Type", "Element type") if column in df.columns),
                    None,
                )
                if type_column:
                    df = df[df[type_column].astype(str).str.upper() == "AMR"]
                filtered_genes = (
                    df[gene_column].replace("", pd.NA).dropna().astype(str).unique()
                )
                gene_list = ", ".join(filtered_genes)
                resistance_dict[sample_id] = gene_list

    return resistance_dict


def read_amrfinderplus_resistance(directory):
    """
    Reads AMRFinderPlus resistance data from multiple TSV files in a given directory.

    Args:
        directory (str): Path to the directory containing AMRFinderPlus result files.

    Returns:
        dict: A dictionary where keys are sample names and values are Pandas DataFrames with resistance data.
    """
    amr_data = {}

    for filename in sorted(os.listdir(directory)):
        if filename.endswith("_out.tsv"):
            sample_id = normalize_sample_id(filename.replace("_out.tsv", ""))
            file_path = os.path.join(directory, filename)

            df = pd.read_csv(file_path, sep="\t", keep_default_na=False)

            # Remove unnecessary columns
            columns_to_remove = [
                "Name",
                "Protein id",
                "Protein identifier",
                "HMM accession",
                "HMM id",
                "HMM description",
            ]
            df_filtered = df.drop(
                columns=[col for col in columns_to_remove if col in df.columns]
            )
            df_filtered = df_filtered.where(pd.notna(df_filtered), None)

            # Store the filtered DataFrame in the dictionary
            amr_data[sample_id] = df_filtered

    return amr_data


def read_variants_stats(variants_stats_file):
    """
    Reads a variant statistics file and extracts SNP, deletion, insertion, and heterozygous mutation counts.

    Args:
        variants_stats_file (str): Path to the variant statistics file.

    Returns:
        dict: A dictionary where keys are sample names and values contain variant statistics.
    """
    try:
        df_variants = pd.read_csv(variants_stats_file, sep=";")
        df_variants.columns = df_variants.columns.str.strip()

        required_columns = ["SAMPLENAME", "SNP", "DEL", "INS", "HET"]
        for col in required_columns:
            if col not in df_variants.columns:
                raise KeyError(f"Missing required column: {col}")

        return {
            str(
                row["SAMPLENAME"]
            ): f"{row['SNP']};{row['DEL']};{row['INS']};{row['HET']}"
            for _, row in df_variants.iterrows()
        }
    except Exception as e:
        print(f"Error processing variants stats file: {e}")
        return {}


def read_quast_report(quast_file):
    """
    Reads a QUAST assembly report and extracts relevant quality metrics.

    Args:
        quast_file (str): Path to the QUAST report file.

    Returns:
        dict: A dictionary where keys are sample names and values contain assembly quality statistics.
    """
    try:
        df_quast = pd.read_csv(quast_file, sep="\t", header=0)
        required_columns = [
            "Assembly",
            "# contigs (>= 1000 bp)",
            "GC (%)",
            "L50",
            "N50",
            "Total length (>= 1000 bp)",
        ]
        for col in required_columns:
            if col not in df_quast.columns:
                raise KeyError(f"Missing required column: {col}")

        df_quast = df_quast.copy()
        df_quast.loc[:, "Sample"] = df_quast["Assembly"].map(normalize_sample_id)

        return df_quast.set_index("Sample")[required_columns].to_dict(orient="index")
    except Exception as e:
        print(f"Error processing QUAST report file: {e}")
        return {}


def read_quast_per_reference(directory):
    """
    Reads all `transposed_report.tsv` files within subdirectories of the given directory.
    Extracts the `# genomic features` and `Genome fraction (%)` values per sample.

    Args:
        directory (str): Path to the QUAST per-reference reports directory.

    Returns:
        dict: A dictionary where keys are sample names (without `.scaffolds`)
              and values are a dictionary with:
              - "genomic_features": Number of genomic features
              - "genome_fraction": Genome fraction (%)
    """
    quast_data = {}

    # Recorrer todos los subdirectorios dentro del directorio dado
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file == "transposed_report.tsv":
                file_path = os.path.join(root, file)

                try:
                    df = pd.read_csv(file_path, sep="\t", header=0)

                    # Verificar que las columnas necesarias existen
                    required_columns = [
                        "Assembly",
                        "# genomic features",
                        "Genome fraction (%)",
                    ]
                    for col in required_columns:
                        if col not in df.columns:
                            raise KeyError(
                                f"Missing required column: {col} in {file_path}"
                            )

                    # Limpiar el nombre de la muestra eliminando `.scaffolds`
                    df["Sample"] = df["Assembly"].map(normalize_sample_id)

                    # Crear diccionario de resultados
                    for _, row in df.iterrows():
                        sample = row["Sample"]
                        quast_data[sample] = {
                            "genomic_features": row["# genomic features"],
                            "genome_fraction": row["Genome fraction (%)"],
                        }

                except Exception as e:
                    print(f"Error processing QUAST report {file_path}: {e}")

    return quast_data


def read_plasmid_data(plasmid_file):
    """
    Reads a plasmid identification report and processes relevant columns.

    Args:
        plasmid_file (str): Path to the plasmid report file.

    Returns:
        DataFrame: A Pandas DataFrame containing plasmid information.
    """
    try:
        df = pd.read_csv(plasmid_file, sep="\t")

        # Column renaming dictionary
        column_rename = {
            "sample": "Sample",
            "id": "Plasmid ID",
            "length": "Length",
            "species": "Species",
            "description": "Description",
            "fraction_covered": "Fraction Covered",
            "contig_number": "Contig Number",
            "% Mapping": "% Mapping",
        }

        # Rename columns for clarity
        df.rename(columns=column_rename, inplace=True)
        return df
    except Exception as e:
        print(f"Error processing plasmid ID file: {e}")
        return {}


def read_mlva_results(mlva_file):
    """
    Reads the MLVA analysis CSV file and extracts relevant columns.

    Args:
        mlva_file (str): Path to the MLVA results file.

    Returns:
        tuple: (header list, dictionary where keys are sample IDs and values are lists of MLVA results).
    """
    try:
        df_mlva = pd.read_csv(mlva_file)

        if "Access_number" not in df_mlva.columns:
            raise KeyError("Missing required column: 'Access_number' in MLVA file")

        mlva_headers = df_mlva.columns[2:].tolist()

        mlva_dict = {
            str(row["Access_number"]): row.iloc[2:].tolist()
            for _, row in df_mlva.iterrows()
        }

        return mlva_headers, mlva_dict

    except Exception as e:
        print(f"Error reading MLVA file: {e}")
        return [], {}


def read_fasta_alignment(file_path):
    """Read an aligned FASTA file while preserving its sequence order."""
    names = []
    sequences = []
    current_name = None
    current_sequence = []

    with open(file_path, "r") as alignment:
        for line_number, raw_line in enumerate(alignment, start=1):
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if current_name is not None:
                    names.append(current_name)
                    sequences.append("".join(current_sequence).upper())
                current_name = line[1:].split()[0]
                if not current_name:
                    raise ValueError(f"Empty FASTA identifier at {file_path}:{line_number}")
                current_sequence = []
            else:
                if current_name is None:
                    raise ValueError(
                        f"Sequence data before the first FASTA header at "
                        f"{file_path}:{line_number}"
                    )
                current_sequence.append(line)

    if current_name is not None:
        names.append(current_name)
        sequences.append("".join(current_sequence).upper())

    if not names:
        raise ValueError(f"No sequences found in alignment: {file_path}")
    if len(names) != len(set(names)):
        raise ValueError(f"Duplicate sequence identifiers in alignment: {file_path}")

    lengths = {len(sequence) for sequence in sequences}
    if len(lengths) != 1:
        raise ValueError(
            f"Sequences have different lengths in alignment {file_path}: "
            f"{sorted(lengths)}"
        )

    return names, sequences


def read_snippy_mapping_reference(commands_file):
    """Read the actual ``--ref`` used by the generated Snippy commands.

    The lablog supplies a wildcard placeholder to ``snippy-multi``; commands.out
    is therefore the authoritative record of the reference used for every run.
    """
    result = {
        "commands_file": commands_file or "Not found",
        "references": [],
        "display": "Not detected",
        "notes": "Snippy commands.out was not found.",
    }
    if not commands_file:
        return result

    references = []
    with open(commands_file, "r") as commands:
        for line in commands:
            try:
                tokens = shlex.split(line)
            except ValueError:
                continue
            if not tokens or os.path.basename(tokens[0]) != "snippy":
                continue
            for index, token in enumerate(tokens):
                if token == "--ref" and index + 1 < len(tokens):
                    references.append(tokens[index + 1])
                elif token.startswith("--ref="):
                    references.append(token.split("=", 1)[1])

    references = list(dict.fromkeys(references))
    result["references"] = references
    if not references:
        result["notes"] = f"No --ref argument found in {commands_file}."
        return result

    descriptions = []
    resolved_files = []
    for reference in references:
        if os.path.isabs(reference):
            candidates = [
                reference,
                re.sub(
                    r"^/scratch/",
                    "/data/ucct/bi/scratch_tmp/",
                    reference,
                ),
            ]
        else:
            candidates = [
                os.path.normpath(os.path.join(os.path.dirname(commands_file), reference)),
                os.path.normpath(os.path.join("../REFERENCES", os.path.basename(reference))),
            ]
        resolved = next((path for path in candidates if os.path.isfile(path)), None)
        resolved_files.append(resolved or reference)

        fasta_header = ""
        if resolved:
            with open(resolved, "r") as fasta:
                for line in fasta:
                    if line.startswith(">"):
                        fasta_header = line[1:].strip()
                        break
        descriptions.append(fasta_header or os.path.basename(reference))

    result["display"] = "; ".join(descriptions)
    result["notes"] = (
        f"Reference file(s): {', '.join(resolved_files)}. "
        f"Detected from --ref in {commands_file}."
    )
    if len(references) > 1:
        result["notes"] += " Warning: multiple distinct references were detected."
    return result


def calculate_pairwise_nucleotide_differences(sequences, chunk_size=100000):
    """Calculate MEGA-style absolute differences with pairwise deletion.

    A position is compared for a pair only when both sequences contain an
    unambiguous A, C, G, or T. Gaps, Ns, and other ambiguity codes are excluded
    independently for each pair.
    """
    sequence_array = np.vstack(
        [np.frombuffer(sequence.encode("ascii"), dtype=np.uint8) for sequence in sequences]
    )
    sequence_count, alignment_length = sequence_array.shape
    differences = np.zeros((sequence_count, sequence_count), dtype=np.int64)
    valid_bases = np.frombuffer(b"ACGT", dtype=np.uint8)

    # Work in position chunks so whole-genome alignments do not require large
    # temporary arrays for every pair.
    for start in range(0, alignment_length, chunk_size):
        block = sequence_array[:, start : start + chunk_size]
        valid = np.isin(block, valid_bases)
        for first in range(sequence_count - 1):
            pair_valid = valid[first + 1 :] & valid[first]
            pair_differences = (block[first + 1 :] != block[first]) & pair_valid
            counts = np.count_nonzero(pair_differences, axis=1)
            differences[first, first + 1 :] += counts

    differences += differences.T
    return differences


# ------------------------------------------------------
# Excel Writing Functions
# ------------------------------------------------------


def add_samples(ws, samples):
    """
    Adds sample names to the first column of the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the samples will be added.
    samples (dict): Dictionary mapping sample names to their respective row numbers.
    """
    try:
        for sample, row_num in samples.items():
            ws[f"A{row_num}"] = sample
    except Exception as e:
        print(f"Error adding samples to xlsx: {e}")
        return {}


def add_ariba_mlst_stats(ws, samples, mlst_dict, mlst_header):
    """
    Adds ARIBA MLST (Multi-Locus Sequence Typing) results to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    mlst_dict (dict): Dictionary containing MLST results for each sample.
    mlst_header (str): Header label for the MLST column.
    """
    try:
        ws["C2"] = mlst_header
        for sample, row_num in samples.items():
            if sample in mlst_dict:
                ws[f"B{row_num}"] = mlst_dict[sample]["ST"]
                ws[f"C{row_num}"] = mlst_dict[sample]["genes"]
    except Exception as e:
        print(f"Error adding ariba mlst to xlsx: {e}")
        return {}


def add_kmerfinder_stats(ws, samples, kmerfinder_dict):
    """
    Adds KmerFinder results to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    kmerfinder_dict (dict): Dictionary containing KmerFinder results.
    """
    try:
        for sample, row_num in samples.items():
            if sample in kmerfinder_dict:
                ws[f"E{row_num}"] = kmerfinder_dict[sample]["colE"]
                ws[f"F{row_num}"] = kmerfinder_dict[sample]["colF"]
                ws[f"G{row_num}"] = kmerfinder_dict[sample]["colG"]
    except Exception as e:
        print(f"Error adding kmerfinder results to xlsx: {e}")
        return {}


def add_mapping_stats(ws, samples, mapping_dict, wgs_metrics_dict):
    """
    Adds mapping statistics to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the statistics will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    mapping_dict (dict): Dictionary containing mapping statistics.
    wgs_metrics_dict (dict): Dictionary containing whole-genome sequencing metrics.
    """
    try:
        for sample, row_num in samples.items():
            ws[f"H{row_num}"] = mapping_dict.get(sample, "NA")
            ws[f"I{row_num}"] = wgs_metrics_dict.get(sample, {}).get(
                "MEAN_COVERAGE", "NA"
            )
            ws[f"J{row_num}"] = wgs_metrics_dict.get(sample, {}).get("PCT_10X", "NA")
    except Exception as e:
        print(f"Error adding mapping stats to xlsx: {e}")
        return {}


def add_variants_stats(ws, samples, variants_dict):
    """
    Adds variant statistics to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the variant data will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    variants_dict (dict): Dictionary containing variant statistics.
    """
    try:
        for sample, row_num in samples.items():
            ws[f"K{row_num}"] = variants_dict.get(sample, "NA;NA;NA;NA")
    except Exception as e:
        print(f"Error adding variant stats to xlsx: {e}")
        return {}


def add_quast_stats(ws, samples, quast_dict):
    """
    Adds QUAST assembly quality statistics to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the QUAST statistics will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    quast_dict (dict): Dictionary containing QUAST results.
    """
    try:
        for sample, row_num in samples.items():
            normalized_sample = normalize_sample_id(sample)
            if normalized_sample in quast_dict:
                quast_stats = quast_dict[normalized_sample]
                ws[f"L{row_num}"] = quast_stats.get("# contigs (>= 1000 bp)", "NA")
                ws[f"N{row_num}"] = quast_stats.get("GC (%)", "NA")
                ws[f"P{row_num}"] = quast_stats.get("L50", "NA")
                ws[f"Q{row_num}"] = quast_stats.get("N50", "NA")
                ws[f"R{row_num}"] = quast_stats.get("Total length (>= 1000 bp)", "NA")
    except Exception as e:
        print(f"Error adding quast stats to xlsx: {e}")
        return {}


def add_quast_per_reference(ws, samples, quast_dict):
    """
    Adds per-reference QUAST results to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the QUAST statistics will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    quast_dict (dict): Dictionary containing QUAST per-reference results.
    """
    try:
        for sample, row_num in samples.items():
            if sample in quast_dict:
                ws[f"M{row_num}"] = quast_dict[sample].get("genomic_features", "NA")
                ws[f"O{row_num}"] = quast_dict[sample].get("genome_fraction", "NA")
    except Exception as e:
        print(f"Error adding per-reference QUAST stats to xlsx: {e}")


def add_virulence_stats(ws, samples, virulence_dict):
    """
    Adds virulence gene statistics to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the virulence statistics will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    virulence_dict (dict): Dictionary containing virulence gene information.
    """
    try:
        for sample, row_num in samples.items():
            if sample in virulence_dict:
                ws[f"A{row_num-1}"] = sample
                ws[f"B{row_num-1}"] = virulence_dict[sample]["genes"]
                ws[f"C{row_num-1}"] = virulence_dict[sample]["count"]
    except Exception as e:
        print(f"Error adding virulence stats to xlsx: {e}")
        return {}


def add_resistance_stats(ws, samples, card_dict):
    """
    Adds antimicrobial resistance (AMR) gene statistics from CARD to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the AMR data will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    card_dict (dict): Dictionary containing AMR gene data.
    """
    try:
        for sample, row_num in samples.items():
            if sample in card_dict:
                ws[f"A{row_num-1}"] = sample
                ws[f"B{row_num-1}"] = card_dict[sample]["genes"]
    except Exception as e:
        print(f"Error adding resistance stats to xlsx: {e}")
        return {}


def add_amrfinder_results(ws, samples, resistance_dict):
    """
    Adds AMRFinder resistance gene results to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the AMRFinder results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    resistance_dict (dict): Dictionary containing AMRFinder resistance gene results.
    """
    try:
        for sample, row_num in samples.items():
            if sample in resistance_dict:
                ws[f"C{row_num-1}"] = resistance_dict[sample]
    except Exception as e:
        print(f"Error adding amrfinder results to xlsx: {e}")
        return {}


def add_amrfinderplus_resistance(ws, samples, amr_data):
    """
    Adds AMRFinderPlus resistance gene results in a tabular format.

    Parameters:
    ws (Worksheet): The Excel worksheet where the AMRFinderPlus results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    amr_data (dict): Dictionary containing AMRFinderPlus resistance gene data as Pandas DataFrames.
    """
    try:
        start_row = 2
        for sample, df in amr_data.items():
            if sample in samples:
                row_num = start_row
                for _, row in df.iterrows():
                    ws[f"A{row_num}"] = sample
                    for col_idx, value in enumerate(row.values, start=2):
                        ws.cell(row=row_num, column=col_idx, value=value)
                    row_num += 1
                start_row = row_num
    except Exception as e:
        print(f"Error adding amrfinder resistance stats to xlsx: {e}")
        return {}


def add_plasmid_data(ws, df):
    """
    Adds plasmid data from a Pandas DataFrame to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the plasmid data will be added.
    df (DataFrame): Pandas DataFrame containing plasmid information.
    """
    try:
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    except Exception as e:
        print(f"Error adding plasmidID stats to xlsx: {e}")
        return {}


def add_mlva_results(ws, mlva_headers, mlva_dict):
    """
    Adds MLVA results to the worksheet, creating new headers dynamically.

    Parameters:
    ws (Worksheet): The Excel worksheet where the MLVA results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    mlva_headers (list): List of MLVA column headers.
    mlva_dict (dict): Dictionary containing MLVA results per sample.
    """
    try:
        for column, header in enumerate(["Sample ID"] + mlva_headers, start=1):
            ws.cell(1, column, header)

        for row_number, (sample_id, values) in enumerate(mlva_dict.items(), start=2):
            for column, value in enumerate([sample_id] + values, start=1):
                ws.cell(row_number, column, value)

    except Exception as e:
        print(f"Error adding MLVA results to xlsx: {e}")


def reset_worksheet(wb, sheet_name):
    """Clear a template worksheet while retaining its name and position."""
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Missing worksheet in template: {sheet_name}")
    index = wb.sheetnames.index(sheet_name)
    wb.remove(wb[sheet_name])
    return wb.create_sheet(sheet_name, index)


def add_readme_sheet(wb, entries):
    """Populate the first worksheet with descriptions and data provenance."""
    ws = reset_worksheet(wb, "README")
    headers = [
        "Sheet",
        "Description",
        "Source",
        "Included samples/sequences",
        "Alignment sites",
        "Notes",
    ]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(1, column, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_number, entry in enumerate(entries, start=2):
        values = [
            entry.get("sheet"),
            entry.get("description"),
            entry.get("source"),
            entry.get("count"),
            entry.get("sites"),
            entry.get("notes"),
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row_number, column, value)
            ws.cell(row_number, column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    widths = [30, 55, 42, 26, 18, 55]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(entries) + 1}"


def apply_homogeneous_formatting(wb):
    """Apply a consistent borderless, filterable layout to every worksheet."""
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    no_border = Border()

    for ws in wb.worksheets:
        header_row = 2 if ws.title == "summary" else 1

        # Determine the real populated range rather than using template-only
        # formatting, which can extend hundreds of blank rows.
        populated_cells = [
            cell
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        if populated_cells:
            last_row = max(cell.row for cell in populated_cells)
            last_column = max(cell.column for cell in populated_cells)
        else:
            last_row = header_row
            last_column = 1

        # Remove borders from every instantiated template cell as well as all
        # newly populated cells.
        for row in ws.iter_rows():
            for cell in row:
                cell.border = no_border

        style_rows = (1, 2) if ws.title == "summary" else (header_row,)
        for row_number in style_rows:
            for column in range(1, last_column + 1):
                cell = ws.cell(row_number, column)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        ws.freeze_panes = f"B{header_row + 1}"
        header_has_values = any(
            ws.cell(header_row, column).value is not None
            for column in range(1, last_column + 1)
        )
        if header_has_values:
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(last_column)}{last_row}"
            )
        else:
            ws.auto_filter.ref = None


def add_snp_distance_sheets(
    wb,
    samples,
    alignment_file,
    matrix_sheet,
    pairs_sheet,
    description,
):
    """Create a symmetric distance matrix and a three-column pair table."""
    ws_matrix = reset_worksheet(wb, matrix_sheet)
    ws_pairs = reset_worksheet(wb, pairs_sheet)

    if not alignment_file:
        print(f"Warning: alignment not found for {description}")
        return {
            "source": "Not found",
            "count": 0,
            "sites": 0,
            "missing": sorted(samples),
        }

    names, sequences = read_fasta_alignment(alignment_file)
    distances = calculate_pairwise_nucleotide_differences(sequences)
    present_samples = {normalize_sample_id(name) for name in names}
    missing_samples = sorted(set(samples) - present_samples)
    site_count = len(sequences[0])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    table_row = 1
    ws_matrix.cell(table_row, 1, "Sample")
    for column, name in enumerate(names, start=2):
        ws_matrix.cell(table_row, column, name)
    for row, name in enumerate(names, start=table_row + 1):
        ws_matrix.cell(row, 1, name)
        for column, value in enumerate(distances[row - table_row - 1], start=2):
            ws_matrix.cell(row, column, int(value))
    for cell in ws_matrix[table_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(text_rotation=90 if cell.column > 1 else 0)
    for row in range(table_row + 1, table_row + 1 + len(names)):
        ws_matrix.cell(row, 1).font = Font(bold=True)
    ws_matrix.freeze_panes = "B2"
    ws_matrix.column_dimensions["A"].width = 28
    for column in range(2, len(names) + 2):
        # Leave enough room for five-digit distances without requiring users
        # to resize every matrix column manually in Excel.
        ws_matrix.column_dimensions[get_column_letter(column)].width = 12

    pair_headers = ["Sample 1", "Sample 2", "Nucleotide differences"]
    for column, header in enumerate(pair_headers, start=1):
        cell = ws_pairs.cell(table_row, column, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    pair_row = table_row + 1
    for first in range(len(names) - 1):
        for second in range(first + 1, len(names)):
            ws_pairs.cell(pair_row, 1, names[first])
            ws_pairs.cell(pair_row, 2, names[second])
            ws_pairs.cell(pair_row, 3, int(distances[first, second]))
            pair_row += 1
    ws_pairs.freeze_panes = "A2"
    ws_pairs.auto_filter.ref = f"A{table_row}:C{pair_row - 1}"
    ws_pairs.column_dimensions["A"].width = 28
    ws_pairs.column_dimensions["B"].width = 28
    ws_pairs.column_dimensions["C"].width = 23

    return {
        "source": alignment_file,
        "count": len(names),
        "sites": site_count,
        "missing": missing_samples,
    }


# ------------------------------------------------------
# Main Functions
# ------------------------------------------------------


def main():
    samples = read_samples(samples_file)

    mlst_dict, mlst_header = read_ariba_mlst(tsv_file) if tsv_file else ({}, "")
    kmerfinder_dict = read_kmerfinder(csv_file) if csv_file else {}
    mapping_dict = read_mapping_stats(mapping_file) if mapping_file else {}
    wgs_metrics_dict = read_wgs_metrics(wgs_metrics_file) if wgs_metrics_file else {}
    variants_dict = (
        read_variants_stats(variants_stats_file) if variants_stats_file else {}
    )
    quast_dict = read_quast_report(quast_report_file) if quast_report_file else {}
    quast_per_reference_dict = read_quast_per_reference(quast_dir) if quast_dir else {}
    virulence_dict = read_virulence_stats(virulence_file) if virulence_file else {}
    resistance_dic = read_card_stats(card_file) if card_file else {}
    amrfinder_dict = read_amrfinder_results(amrfinder_dir) if amrfinder_dir else {}
    amr_resistance_data = (
        read_amrfinderplus_resistance(amrfinder_dir) if amrfinder_dir else {}
    )
    plasmid_data = read_plasmid_data(plasmid_file) if plasmid_file else pd.DataFrame()
    mlva_headers, mlva_dict = read_mlva_results(mlva_file) if mlva_file else ([], {})
    mapping_reference = read_snippy_mapping_reference(snippy_commands_file)

    wb = load_workbook(xlsx_template)

    ws_summary = wb["summary"]
    ws_plasmids = wb["plasmids"]
    ws_virulence = wb["virulence"]
    ws_resistance = wb["Resistance result"]
    ws_mlva = wb["MLVA"]
    ws_amrfinder_resistance = wb["AMRFinderPlus Resistance result"]

    if mapping_reference["display"] == "Not detected":
        ws_summary["H1"] = "MAPPING: reference not detected"
    else:
        ws_summary["H1"] = f"MAPPING: {mapping_reference['display']}"

    add_samples(ws_summary, samples)
    add_ariba_mlst_stats(ws_summary, samples, mlst_dict, mlst_header)
    add_kmerfinder_stats(ws_summary, samples, kmerfinder_dict)
    add_mapping_stats(ws_summary, samples, mapping_dict, wgs_metrics_dict)
    add_variants_stats(ws_summary, samples, variants_dict)
    add_quast_stats(ws_summary, samples, quast_dict)
    add_quast_per_reference(ws_summary, samples, quast_per_reference_dict)
    add_virulence_stats(ws_virulence, samples, virulence_dict)
    add_resistance_stats(ws_resistance, samples, resistance_dic)
    add_amrfinder_results(ws_resistance, samples, amrfinder_dict)
    add_amrfinderplus_resistance(ws_amrfinder_resistance, samples, amr_resistance_data)

    core_all_sites = add_snp_distance_sheets(
        wb,
        samples,
        phylo_alignment_file,
        "SNP core all sites",
        "SNP core all sites pairs",
        "Core-genome invariant and variable positions called in all included sequences",
    )
    variable_no_recomb = add_snp_distance_sheets(
        wb,
        samples,
        nonrecombinant_alignment_file,
        "SNP variable no recomb",
        "SNP variable no recomb pairs",
        "Variable positions after removal of recombinant regions",
    )

    if not plasmid_data.empty:
        add_plasmid_data(ws_plasmids, plasmid_data)

    if mlva_headers and mlva_dict:
        add_mlva_results(ws_mlva, mlva_headers, mlva_dict)
    else:
        ws_mlva["A1"] = "Sample ID"

    distance_method = (
        "Absolute nucleotide differences with pairwise deletion; only A/C/G/T "
        "positions are compared, and gaps or ambiguous bases are excluded per pair."
    )

    def distance_notes(metadata):
        missing = metadata.get("missing", [])
        missing_note = (
            f" Absent outbreak samples: {', '.join(missing)}."
            if missing
            else " All outbreak samples are present."
        )
        return distance_method + missing_note

    plasmid_count = (
        plasmid_data["Sample"].astype(str).nunique()
        if not plasmid_data.empty and "Sample" in plasmid_data.columns
        else 0
    )
    readme_entries = [
        {
            "sheet": "README",
            "description": "Workbook contents, sources, counts, and calculation notes.",
        },
        {
            "sheet": "summary",
            "description": "Per-sample typing, taxonomy, mapping, variants, and assembly statistics.",
            "source": f"Analysis summary files; {mapping_reference['commands_file']}",
            "count": len(samples),
            "notes": (
                f"Snippy mapping reference: {mapping_reference['display']}. "
                f"{mapping_reference['notes']}"
            ),
        },
        {
            "sheet": "SNP core all sites",
            "description": "Symmetric SNP-distance matrix across invariant and variable core-genome positions called in all included sequences.",
            "source": core_all_sites["source"],
            "count": core_all_sites["count"],
            "sites": core_all_sites["sites"],
            "notes": distance_notes(core_all_sites),
        },
        {
            "sheet": "SNP core all sites pairs",
            "description": "Long-format pair list corresponding to the SNP core all sites matrix.",
            "source": core_all_sites["source"],
            "count": core_all_sites["count"],
            "sites": core_all_sites["sites"],
            "notes": distance_notes(core_all_sites),
        },
        {
            "sheet": "SNP variable no recomb",
            "description": "Symmetric SNP-distance matrix across variable positions after recombinant regions were removed.",
            "source": variable_no_recomb["source"],
            "count": variable_no_recomb["count"],
            "sites": variable_no_recomb["sites"],
            "notes": distance_notes(variable_no_recomb),
        },
        {
            "sheet": "SNP variable no recomb pairs",
            "description": "Long-format pair list corresponding to the variable non-recombinant matrix.",
            "source": variable_no_recomb["source"],
            "count": variable_no_recomb["count"],
            "sites": variable_no_recomb["sites"],
            "notes": distance_notes(variable_no_recomb),
        },
        {
            "sheet": "plasmids",
            "description": "Per-sample PlasmidID matches.",
            "source": plasmid_file or "Not found",
            "count": plasmid_count,
        },
        {
            "sheet": "virulence",
            "description": "Per-sample virulence genes detected with ARIBA/VFDB.",
            "source": virulence_file or "Not found",
            "count": len(virulence_dict),
        },
        {
            "sheet": "Resistance result",
            "description": "Per-sample antimicrobial-resistance gene summary from CARD and AMRFinderPlus.",
            "source": "ARIBA/CARD and AMRFinderPlus",
            "count": len(set(resistance_dic) | set(amrfinder_dict)),
        },
        {
            "sheet": "AMRFinderPlus Resistance result",
            "description": "Detailed AMRFinderPlus element calls.",
            "source": amrfinder_dir or "Not found",
            "count": len(amr_resistance_data),
        },
        {
            "sheet": "MLVA",
            "description": "Multiple-locus variable-number tandem-repeat analysis results.",
            "source": mlva_file or "Not found",
            "count": len(mlva_dict),
        },
    ]
    add_readme_sheet(wb, readme_entries)
    apply_homogeneous_formatting(wb)

    output_xlsx = "summary_outbreak_filled.xlsx"
    wb.save(output_xlsx)
    print(f"File saved: {output_xlsx}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Find closely related Snippy samples and report their differing SNPs.

Distances are calculated directly from snippy-core's core.tab.  For each pair
below the requested threshold, the script looks up supporting read counts in
each sample's snps.vcf (or snps.vcf.gz). Reference calls, which are absent
from variant-only VCFs, are measured from snps.bam with samtools mpileup.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import itertools
import json
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

VALID_BASES = frozenset("ACGT")


def parse_args() -> argparse.Namespace:
    """Parse command-line options for the close-pair SNP QC workflow."""
    parser = argparse.ArgumentParser(
        description=(
            "Calculate pairwise SNP distances from core.tab and extract VCF "
            "evidence for variants in pairs with fewer than N differences."
        )
    )
    parser.add_argument(
        "snippy_dir",
        nargs="?",
        type=Path,
        default=Path("."),
        help="snippy-core output directory (default: current directory)",
    )
    parser.add_argument(
        "--core-tab",
        type=Path,
        help="core.tab path (default: SNIPPY_DIR/core.tab)",
    )
    parser.add_argument(
        "--threshold",
        type=int,
        default=20,
        help="Retain pairs with fewer than this many SNPs (default: 20)",
    )
    parser.add_argument(
        "--output-prefix",
        type=Path,
        default=Path("close_pair_snp_qc"),
        help="Output file prefix (default: close_pair_snp_qc)",
    )
    parser.add_argument(
        "--matrices-dir",
        type=Path,
        help="Directory containing generated SNP-distance matrices (default: SNIPPY_DIR)",
    )
    parser.add_argument(
        "--gubbins-gff",
        type=Path,
        help="Gubbins recombination GFF (default: SNIPPY_DIR/gubbins.recombination_predictions.gff)",
    )
    parser.add_argument(
        "--missing-vcf",
        choices=("error", "warn"),
        default="error",
        help="Action to perform when a selected sample has no VCF (default: error)",
    )
    parser.add_argument(
        "--reference",
        type=Path,
        help="Reference FASTA override for mpileup (default: each sample's reference/ref.fa)",
    )
    parser.add_argument(
        "--samtools",
        default="samtools",
        help="Samtools executable or path (default: samtools)",
    )
    parser.add_argument(
        "--no-bam-evidence",
        action="store_true",
        help="Do not obtain evidence for reference calls from snps.bam",
    )
    parser.add_argument(
        "--pileup-mapq",
        type=int,
        default=10,
        help="Minimum mapping quality for mpileup evidence (default: 10)",
    )
    parser.add_argument(
        "--pileup-baseq",
        type=int,
        default=5,
        help="Minimum base quality for mpileup evidence (default: 5)",
    )
    parser.add_argument(
        "--warn-min-depth",
        type=float,
        default=10,
        help="Warn when evidence depth is below this value (default: 10)",
    )
    parser.add_argument(
        "--warn-min-called-fraction",
        type=float,
        default=0.90,
        help="Warn when the called allele fraction is below this value (default: 0.90)",
    )
    parser.add_argument(
        "--no-qc-warnings",
        action="store_true",
        help="Suppress per-site QC warnings on standard output",
    )
    parser.add_argument(
        "--warn-min-strand-fraction",
        type=float,
        default=0.10,
        help="Warn if less than this fraction of called reads is on either strand (default: 0.10)",
    )
    parser.add_argument(
        "--warn-indel-distance",
        type=int,
        default=5,
        help="Warn for SNPs within these many bases of a raw-VCF indel/complex call (default: 5)",
    )
    parser.add_argument(
        "--warn-homopolymer-run",
        type=int,
        default=5,
        help="Warn when FreeBayes' RUN field is at least this length (default: 5)",
    )
    parser.add_argument(
        "--cluster-window",
        type=int,
        default=100,
        help="Window size in bases for local pairwise SNP clusters (default: 100)",
    )
    parser.add_argument(
        "--cluster-min-snps",
        type=int,
        default=3,
        help="Minimum pairwise SNPs in the window to flag a cluster (default: 3)",
    )
    parser.add_argument(
        "--no-summary",
        action="store_true",
        help="Do not write the human-readable variant summary table",
    )
    parser.add_argument(
        "--no-metadata",
        action="store_true",
        help="Do not write the JSON run metadata and provenance file",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Do not write the Excel report with README and result sheets",
    )
    parser.add_argument(
        "--excel-output",
        type=Path,
        help="Excel report filename (default: OUTPUT_PREFIX_report.xlsx)",
    )
    return parser.parse_args()


def read_core_tab(
    path: Path,
) -> tuple[list[str], list[tuple[str, int, str, list[str]]]]:
    """Read core.tab without discarding non-canonical calls.

    Calls are upper-cased.  Pairwise A/C/G/T filtering is deferred to the following
    ``find_close_pairs`` function so that excluded positions can be reported.
    """
    with path.open(newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"Empty core table: {path}")
        if len(header) < 4 or header[:3] != ["CHR", "POS", "REF"]:
            raise ValueError(f"Unexpected core.tab header in {path}")
        samples = header[3:]
        if len(samples) != len(set(samples)):
            raise ValueError("Duplicate sample names in core.tab")
        sites = []
        for line_number, row in enumerate(reader, 2):
            if not row:
                continue
            if len(row) != len(header):
                raise ValueError(
                    f"Line {line_number} has {len(row)} columns; expected {len(header)}"
                )
            sites.append(
                (row[0], int(row[1]), row[2].upper(), [x.upper() for x in row[3:]])
            )
    return samples, sites


def find_close_pairs(samples, sites, threshold):
    """Find pairs with strictly fewer than specified ``threshold`` comparable SNPs.

    A site is comparable only when both calls are A/C/G/T; all other calls are
    pairwise-deleted rather than treated as a match or a difference.
    """
    results = []
    for left, right in itertools.combinations(range(len(samples)), 2):
        differences = []
        comparable = 0
        ignored = 0
        for chrom, pos, ref, calls in sites:
            a, b = calls[left], calls[right]
            if a in VALID_BASES and b in VALID_BASES:
                comparable += 1
                if a != b:
                    differences.append((chrom, pos, ref, a, b))
            else:
                ignored += 1
        if len(differences) < threshold:
            results.append(
                {
                    "sample_1": samples[left],
                    "sample_2": samples[right],
                    "snp_distance": len(differences),
                    "comparable_core_sites": comparable,
                    "ignored_non_acgt_sites": ignored,
                    "differences": differences,
                }
            )
    return sorted(
        results, key=lambda x: (x["snp_distance"], x["sample_1"], x["sample_2"])
    )


def read_distance_matrix(path: Path) -> dict[frozenset[str], int]:
    """Read a symmetric TSV matrix and validate matching row/column order."""
    with path.open(newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"Empty distance matrix: {path}")
        names = header[1:]
        if not names or len(names) != len(set(names)):
            raise ValueError(f"Invalid or duplicate matrix sample names: {path}")
        distances = {}
        observed_rows = []
        for row_number, row in enumerate(reader, 2):
            if len(row) != len(header):
                raise ValueError(f"Wrong column count at {path}:{row_number}")
            row_name = row[0]
            observed_rows.append(row_name)
            for column, value in enumerate(row[1:]):
                if row_name != names[column]:
                    distances[frozenset((row_name, names[column]))] = int(value)
        if observed_rows != names:
            raise ValueError(f"Matrix row and column sample order differs: {path}")
    return distances


def read_gubbins_intervals(path: Path) -> list[tuple[int, int]]:
    """Read and merge 1-based inclusive recombinant intervals from Gubbins."""
    intervals = []
    with path.open() as handle:
        for line_number, line in enumerate(handle, 1):
            if not line.strip() or line.startswith("#"):
                continue
            fields = line.rstrip("\n").split("\t")
            if len(fields) < 5:
                raise ValueError(f"Malformed Gubbins GFF at {path}:{line_number}")
            intervals.append((int(fields[3]), int(fields[4])))
    merged = []
    for start, end in sorted(intervals):
        if merged and start <= merged[-1][1] + 1:
            merged[-1] = (merged[-1][0], max(merged[-1][1], end))
        else:
            merged.append((start, end))
    return merged


def position_in_intervals(position: int, intervals: list[tuple[int, int]]) -> bool:
    """Return whether a 1-based position lies in a merged inclusive interval"""
    import bisect

    starts = [start for start, _ in intervals]
    index = bisect.bisect_right(starts, position) - 1
    return index >= 0 and position <= intervals[index][1]


def open_text(path: Path):
    """Open a plain-text or gzip-compressed file for text reading"""
    return gzip.open(path, "rt") if path.suffix == ".gz" else path.open()


def parse_list(value: str, cast=int):
    """Parse a comma-separated VCF value, retaining invalid items as ``None``."""
    if value in ("", "."):
        return []
    result = []
    for item in value.split(","):
        try:
            result.append(cast(item))
        except ValueError:
            result.append(None)
    return result


def scalar_or_alt(value: str | None, alt_index: int):
    """Select the value corresponding to one ALT allele from a VCF field"""
    if value in (None, "", "."):
        return ""
    values = value.split(",")
    return values[alt_index] if alt_index < len(values) else value


def read_vcf(path: Path) -> dict[tuple[str, int, str], dict[str, str]]:
    """Index SNP evidence by (chromosome, position, called base).

    Equal-length MNP/complex substitutions are expanded base-by-base so that
    sites emitted separately in core.tab still find their source VCF record.
    """
    index = {}
    with open_text(path) as handle:
        for line in handle:
            if not line or line.startswith("#"):
                continue
            fields = line.rstrip("\n").split("\t")
            if len(fields) < 8:
                continue
            chrom, pos_text, _, ref, alt_text, qual, filt, info_text = fields[:8]
            pos = int(pos_text)
            info = {}
            for entry in info_text.split(";"):
                key, sep, value = entry.partition("=")
                if sep:
                    info[key] = value
            fmt = {}
            if len(fields) >= 10:
                fmt = dict(zip(fields[8].split(":"), fields[9].split(":")))
            dp = fmt.get("DP", info.get("DP", ""))
            ro = fmt.get("RO", info.get("RO", ""))
            ao_all = fmt.get("AO", info.get("AO", ""))
            for alt_index, alt in enumerate(alt_text.split(",")):
                evidence = {
                    "vcf_record_pos": pos_text,
                    "vcf_ref": ref,
                    "vcf_alt": alt,
                    "qual": qual if qual != "." else "",
                    "filter": filt,
                    "genotype": fmt.get("GT", ""),
                    "depth": dp,
                    "ref_count": ro,
                    "alt_count": scalar_or_alt(ao_all, alt_index),
                    "variant_type": scalar_or_alt(info.get("TYPE"), alt_index),
                    "ref_forward_count": info.get("SRF", ""),
                    "ref_reverse_count": info.get("SRR", ""),
                    "alt_forward_count": scalar_or_alt(info.get("SAF"), alt_index),
                    "alt_reverse_count": scalar_or_alt(info.get("SAR"), alt_index),
                    "homopolymer_run": info.get("RUN", ""),
                    "nearest_indel_distance": "",
                }
                if len(ref) == len(alt):
                    for offset, (ref_base, alt_base) in enumerate(
                        zip(ref.upper(), alt.upper())
                    ):
                        if ref_base != alt_base and alt_base in VALID_BASES:
                            index[(chrom, pos + offset, alt_base)] = evidence
                elif len(ref) == 1 and len(alt) == 1 and alt.upper() in VALID_BASES:
                    index[(chrom, pos, alt.upper())] = evidence
    return index


def locate_vcf(snippy_dir: Path, sample: str) -> Path | None:
    """Locate a sample's filtered VCF, accepting plain or gzip-compressed files."""
    for name in ("snps.vcf", "snps.vcf.gz"):
        candidate = snippy_dir / sample / name
        if candidate.is_file():
            return candidate
    return None


def locate_raw_vcf(snippy_dir: Path, sample: str) -> Path | None:
    """Locate a sample's raw VCF, accepting plain or gzip-compressed files."""
    for name in ("snps.raw.vcf", "snps.raw.vcf.gz"):
        candidate = snippy_dir / sample / name
        if candidate.is_file():
            return candidate
    return None


def read_non_snp_positions(path: Path) -> dict[str, list[int]]:
    """Find raw-VCF positions whose INFO/TYPE includes a non-SNP event."""
    positions: dict[str, list[int]] = {}
    with open_text(path) as handle:
        for line in handle:
            if line.startswith("#"):
                continue
            fields = line.rstrip("\n").split("\t")
            if len(fields) < 8:
                continue
            info = dict(
                entry.split("=", 1) for entry in fields[7].split(";") if "=" in entry
            )
            types = info.get("TYPE", "").split(",")
            if any(kind and kind != "snp" for kind in types):
                positions.setdefault(fields[0], []).append(int(fields[1]))
    for values in positions.values():
        values.sort()
    return positions


def nearest_distance(sorted_positions: list[int], position: int) -> int | None:
    """Return the distance to the closest position in a sorted coordinate list."""
    import bisect

    index = bisect.bisect_left(sorted_positions, position)
    candidates = sorted_positions[max(0, index - 1) : index + 1]
    return min((abs(position - other) for other in candidates), default=None)


def parse_pileup_bases(text: str, reference: str) -> dict[str, dict[str, int]]:
    """Count A/C/G/T observations in a samtools mpileup read-bases field."""
    counts = {base: {"forward": 0, "reverse": 0} for base in VALID_BASES}
    i = 0
    while i < len(text):
        symbol = text[i]
        if symbol == "^":  # start of read plus one mapping-quality character
            i += 2
        elif symbol == "$":
            i += 1
        elif symbol in "+-":  # indel length and inserted/deleted sequence
            i += 1
            start = i
            while i < len(text) and text[i].isdigit():
                i += 1
            length = int(text[start:i]) if i > start else 0
            i += length
        elif symbol in ".,":
            if reference in counts:
                strand = "forward" if symbol == "." else "reverse"
                counts[reference][strand] += 1
            i += 1
        elif symbol.upper() in counts:
            strand = "forward" if symbol.isupper() else "reverse"
            counts[symbol.upper()][strand] += 1
            i += 1
        else:  # N/n, deletion placeholders (* or #), and reference skips
            i += 1
    return counts


def read_bam_pileup(
    samtools: str,
    reference: Path,
    bam: Path,
    positions: set[tuple[str, int]],
    minimum_mapq: int,
    minimum_baseq: int,
) -> dict[tuple[str, int], dict[str, object]]:
    """Run ``samtools mpileup`` for selected 1-based sites.

    Sites are written as 0-based half-open BED intervals. Mapping quality
    (``-q``) and base quality (``-Q``) are applied before parsing A/C/G/T
    support and strand orientation from the read-bases field.
    """
    if not positions:
        return {}
    with tempfile.NamedTemporaryFile("w", suffix=".bed") as bed:
        for chrom, pos in sorted(positions):
            bed.write(f"{chrom}\t{pos - 1}\t{pos}\n")
        bed.flush()
        command = [
            samtools,
            "mpileup",
            "-q",
            str(minimum_mapq),
            "-Q",
            str(minimum_baseq),
            "-f",
            str(reference),
            "-l",
            bed.name,
            str(bam),
        ]
        result = subprocess.run(command, text=True, capture_output=True)
    if result.returncode:
        detail = result.stderr.strip() or "unknown samtools error"
        raise RuntimeError(f"samtools mpileup failed for {bam}: {detail}")
    index = {}
    for line in result.stdout.splitlines():
        fields = line.split("\t")
        if len(fields) < 5:
            continue
        chrom, pos_text, reference_base, depth, read_bases = fields[:5]
        reference_base = reference_base.upper()
        index[(chrom, int(pos_text))] = {
            "depth": depth,
            "counts": parse_pileup_bases(read_bases, reference_base),
        }
    return index


def evidence_for(index, bam_index, chrom, pos, core_ref, call, other_call):
    """Select VCF evidence, BAM evidence for a reference call, or a status."""
    record = index.get((chrom, pos, call))
    if record is not None:
        return "variant_record", record
    if call == core_ref and (chrom, pos) in bam_index:
        pileup = bam_index[(chrom, pos)]
        counts = pileup["counts"]
        ref_strands = counts.get(core_ref, {"forward": 0, "reverse": 0})
        alt_strands = counts.get(other_call, {"forward": 0, "reverse": 0})
        evidence = {key: "" for key in EVIDENCE_COLUMNS}
        evidence.update(
            {
                "vcf_ref": core_ref,
                "vcf_alt": other_call,
                "depth": pileup["depth"],
                "ref_count": sum(ref_strands.values()),
                "alt_count": sum(alt_strands.values()),
                "variant_type": "reference_call",
                "ref_forward_count": ref_strands["forward"],
                "ref_reverse_count": ref_strands["reverse"],
                "alt_forward_count": alt_strands["forward"],
                "alt_reverse_count": alt_strands["reverse"],
                "nearest_indel_distance": pileup.get("nearest_indel_distance", ""),
            }
        )
        return "reference_call_mpileup", evidence
    status = (
        "reference_call_no_bam_evidence"
        if call == core_ref
        else "variant_not_found_in_vcf"
    )
    return status, {key: "" for key in EVIDENCE_COLUMNS}


EVIDENCE_COLUMNS = (
    "vcf_record_pos",
    "vcf_ref",
    "vcf_alt",
    "qual",
    "filter",
    "genotype",
    "depth",
    "ref_count",
    "alt_count",
    "variant_type",
    "ref_forward_count",
    "ref_reverse_count",
    "alt_forward_count",
    "alt_reverse_count",
    "homopolymer_run",
    "nearest_indel_distance",
)


def add_fractions(row: dict[str, object], prefix: str) -> None:
    """Calculate reference/alternate support fractions using total DP."""
    try:
        dp = float(row[f"{prefix}_depth"])
    except (ValueError, TypeError):
        dp = 0
    for count_name in ("ref", "alt"):
        try:
            count = float(row[f"{prefix}_{count_name}_count"])
            value = count / dp if dp else ""
        except (ValueError, TypeError):
            value = ""
        row[f"{prefix}_{count_name}_fraction"] = value


def numeric(value: object) -> float | None:
    """Convert an evidence value to ``float``, or return ``None`` if unavailable."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def collect_qc_warnings(
    row: dict[str, object],
    sample_number: int,
    minimum_depth: float,
    minimum_called_fraction: float,
    minimum_strand_fraction: float,
    maximum_indel_distance: int,
    minimum_homopolymer_run: int,
) -> list[tuple[str, str]]:
    """Return QC warnings for one sample at one SNP site."""
    prefix = f"sample_{sample_number}"
    sample = str(row[prefix])
    location = f"{row['chrom']}:{row['pos']}"
    context = (
        f"sample={sample} position={location} call={row[f'{prefix}_call']} "
        f"pair={row['pair_id']}"
    )
    status = row[f"{prefix}_evidence_status"]
    if status not in ("variant_record", "reference_call_mpileup"):
        return [("missing_evidence", f"{context} status={status}")]

    warnings = []
    depth = numeric(row[f"{prefix}_depth"])
    if depth is None:
        warnings.append(("missing_depth", context))
    elif depth < minimum_depth:
        warnings.append(
            ("low_depth", f"{context} depth={depth:g} threshold={minimum_depth:g}")
        )

    fraction_name = (
        "ref_fraction" if row[f"{prefix}_call"] == row["core_ref"] else "alt_fraction"
    )
    called_fraction = numeric(row[f"{prefix}_{fraction_name}"])
    if called_fraction is None:
        warnings.append(("missing_called_fraction", context))
    elif called_fraction < minimum_called_fraction:
        warnings.append(
            (
                "low_called_fraction",
                f"{context} called_fraction={called_fraction:.4f} "
                f"threshold={minimum_called_fraction:.4f}",
            )
        )

    called_prefix = "ref" if row[f"{prefix}_call"] == row["core_ref"] else "alt"
    forward = numeric(row[f"{prefix}_{called_prefix}_forward_count"])
    reverse = numeric(row[f"{prefix}_{called_prefix}_reverse_count"])
    if forward is not None and reverse is not None and forward + reverse > 0:
        minor_fraction = min(forward, reverse) / (forward + reverse)
        if minor_fraction < minimum_strand_fraction:
            warnings.append(
                (
                    "strand_bias",
                    f"{context} forward={forward:g} reverse={reverse:g} "
                    f"minor_strand_fraction={minor_fraction:.4f} "
                    f"threshold={minimum_strand_fraction:.4f}",
                )
            )

    indel_distance = numeric(row[f"{prefix}_nearest_indel_distance"])
    if indel_distance is not None and indel_distance <= maximum_indel_distance:
        warnings.append(
            (
                "near_indel_or_complex",
                f"{context} distance={indel_distance:g} threshold={maximum_indel_distance}",
            )
        )
    run = numeric(row[f"{prefix}_homopolymer_run"])
    if run is not None and run >= minimum_homopolymer_run:
        warnings.append(
            (
                "homopolymer_context",
                f"{context} run_length={run:g} threshold={minimum_homopolymer_run}",
            )
        )
    return warnings



def as_float(value: object) -> float | None:
    """Return a numeric value, or ``None`` when an evidence field is absent."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def format_number(value: object) -> str:
    """Format an evidence count without adding a misleading decimal point."""
    numeric_value = as_float(value)
    if numeric_value is None:
        return "not_available"
    return f"{numeric_value:g}"


def format_fraction(value: object) -> str:
    """Format a proportion for the human-readable evidence summary."""
    numeric_value = as_float(value)
    return "not_available" if numeric_value is None else f"{numeric_value:.1%}"


def annotate_variant_row(row: dict[str, str]) -> dict[str, str]:
    """Add interpretation-ready evidence fields without changing raw evidence"""
    for number in (1, 2):
        prefix = f"sample_{number}"
        call = row[f"{prefix}_call"]
        called_kind = "ref" if call == row["core_ref"] else "alt"
        status = row.get(f"{prefix}_evidence_status", "")
        count = row.get(f"{prefix}_{called_kind}_count", "")
        fraction = row.get(f"{prefix}_{called_kind}_fraction", "")
        forward = row.get(f"{prefix}_{called_kind}_forward_count", "")
        reverse = row.get(f"{prefix}_{called_kind}_reverse_count", "")
        forward_value, reverse_value = as_float(forward), as_float(reverse)
        if forward_value is None or reverse_value is None or forward_value + reverse_value == 0:
            minor_fraction = ""
        else:
            minor_fraction = str(min(forward_value, reverse_value) / (forward_value + reverse_value))
        evidence_available = status in ("variant_record", "reference_call_mpileup")
        row[f"{prefix}_called_allele_minor_strand_fraction"] = minor_fraction
        if not evidence_available:
            row[f"{prefix}_evidence_summary"] = f"{call}: evidence not available ({status})"
        else:
            row[f"{prefix}_evidence_summary"] = (
                f"{call}: {format_number(count)}/{format_number(row.get(f'{prefix}_depth', ''))} "
                f"({format_fraction(fraction)}); strands {format_number(forward)}/{format_number(reverse)}"
            )
    warnings = row.get("qc_warnings", "")
    row["qc_status"] = "PASS" if not warnings else "REVIEW"
    row["review_reasons"] = warnings or "none"
    return row


def enhance_output_tables(pairs_path: Path, variants_path: Path, threshold: int) -> list[dict[str, str]]:
    """Include additional useful interpretation fields to the TSV output files"""
    pair_extra = [
        "pair_selection_rule", "comparable_core_fraction", "matrix_availability",
        "pair_qc_status", "pair_review_reasons",
    ]
    with pairs_path.open(newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        pair_rows = list(reader)
        pair_fields = list(reader.fieldnames or [])
    for row in pair_rows:
        comparable = as_float(row.get("comparable_core_sites")) or 0
        ignored = as_float(row.get("ignored_non_acgt_sites")) or 0
        total = comparable + ignored
        row["pair_selection_rule"] = f"snp_distance < {threshold}"
        row["comparable_core_fraction"] = "" if total == 0 else str(comparable / total)
        missing = [
            label for label, field in (
                ("core_tab", "core_tab_matrix_distance"),
                ("phylo", "phylo_aln_distance"),
                ("clean_core", "clean_core_aln_distance"),
            ) if not row.get(field, "")
        ]
        row["matrix_availability"] = "all_available" if not missing else "missing:" + ",".join(missing)
        warnings = row.get("distance_comparison_warnings", "")
        row["pair_qc_status"] = "PASS" if not warnings else "REVIEW"
        row["pair_review_reasons"] = warnings or "none"
    with tempfile.NamedTemporaryFile("w", newline="", dir=pairs_path.parent, delete=False) as temporary:
        writer = csv.DictWriter(temporary, fieldnames=pair_fields + pair_extra, delimiter="\t")
        writer.writeheader()
        writer.writerows(pair_rows)
        temporary_path = Path(temporary.name)
    temporary_path.replace(pairs_path)

    with variants_path.open(newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        variant_rows = [annotate_variant_row(dict(row)) for row in reader]
        variant_fields = list(reader.fieldnames or [])
    derived_fields = [
        field for number in (1, 2) for field in (
            f"sample_{number}_called_allele_minor_strand_fraction",
            f"sample_{number}_evidence_summary",
        )
    ] + ["qc_status", "review_reasons"]
    context_fields = [
        field for field in variant_fields
        if (not field.startswith("sample_") or field in ("sample_1", "sample_2"))
        and field != "qc_warnings"
    ]
    sample_1_fields = [field for field in variant_fields + derived_fields if field.startswith("sample_1_")]
    sample_2_fields = [field for field in variant_fields + derived_fields if field.startswith("sample_2_")]
    qc_fields = ["qc_warnings", "qc_status", "review_reasons"]
    ordered_variant_fields = context_fields + sample_1_fields + sample_2_fields + qc_fields
    with tempfile.NamedTemporaryFile("w", newline="", dir=variants_path.parent, delete=False) as temporary:
        writer = csv.DictWriter(temporary, fieldnames=ordered_variant_fields, delimiter="\t")
        writer.writeheader()
        writer.writerows(variant_rows)
        temporary_path = Path(temporary.name)
    temporary_path.replace(variants_path)
    return variant_rows


def write_variant_summary(path: Path, rows: list[dict[str, str]]) -> None:
    """Write one concise, report-oriented row per pair-specific SNP"""
    fields = [
        "pair_id", "sample_1", "sample_2", "snp_distance", "chrom", "pos", "core_ref",
        "sample_1_call", "sample_2_call", "sample_1_evidence_summary",
        "sample_2_evidence_summary", "removed_by_gubbins", "pair_clustered_snp",
        "qc_status", "review_reasons",
    ]
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def column_description(column: str) -> tuple[str, str]:
    """
    Return a detailed source and definition explanation for every output field of the TSV output files,
    to make column interpretation easier.
    """
    definitions = {
        "pair_id": ("Generated as pair_0001, pair_0002, … in ascending distance order.", "Identifier linking a pair row to its SNP-variant rows."),
        "sample_1": ("Sample-name column in core.tab.", "First sample in the pair."),
        "sample_2": ("Sample-name column in core.tab.", "Second sample in the pair."),
        "snp_distance": ("Calculated from core.tab calls.", "Number of different A/C/G/T calls between the two samples."),
        "comparable_core_sites": ("Calculated from core.tab calls.", "Positions where both samples have A, C, G or T and can be compared."),
        "ignored_non_acgt_sites": ("Calculated from core.tab calls.", "Positions excluded because at least one call is not A, C, G or T."),
        "core_tab_matrix_distance": ("core_tab_snp_distances.tsv: matrix cell for this pair.", "Distance independently calculated from core.tab, when the matrix is available."),
        "core_tab_matrix_matches": ("Comparison of snp_distance and core_tab_matrix_distance.", "yes when both distances agree; no when they differ."),
        "phylo_aln_distance": ("phylo_snp_distances.tsv: matrix cell for this pair.", "Distance in the phylogenetic alignment, when available."),
        "phylo_matches_core_tab": ("Comparison of snp_distance and phylo_aln_distance.", "yes when both distances agree; no when they differ."),
        "clean_core_aln_distance": ("clean_core_snp_distances.tsv: matrix cell for this pair.", "Distance after the clean/Gubbins-associated alignment, when available."),
        "snps_removed_by_gubbins": ("snp_distance − clean_core_aln_distance.", "Differences no longer contributing to the clean-alignment distance."),
        "clean_distance_valid": ("Comparison of clean_core_aln_distance and snp_distance.", "yes when the clean distance is not greater than the original distance."),
        "distance_comparison_warnings": ("Calculated from the three distance-matrix comparisons.", "Semicolon-separated matrix discrepancies or missing-matrix warnings."),
        "pair_selection_rule": ("Command-line parameter --threshold.", "Strict rule used to retain this pair."),
        "comparable_core_fraction": ("comparable_core_sites / (comparable_core_sites + ignored_non_acgt_sites).", "Fraction of core.tab positions usable for this pair."),
        "matrix_availability": ("Presence of the three generated distance-matrix files.", "States whether core.tab, phylo and clean-core comparison matrices were available."),
        "pair_qc_status": ("distance_comparison_warnings.", "PASS when no pair-level matrix warning exists; REVIEW otherwise."),
        "pair_review_reasons": ("distance_comparison_warnings.", "Pair-level warnings, or none."),
        "chrom": ("core.tab: CHR.", "Chromosome identifier."),
        "pos": ("core.tab: POS.", "Genomic coordinate."),
        "core_ref": ("core.tab: REF.", "Reference base in core.tab."),
        "sample_1_call": ("core.tab: sample_1 column at this position.", "Base called for sample_1."),
        "sample_2_call": ("core.tab: sample_2 column at this position.", "Base called for sample_2."),
        "pair_local_snp_count": ("Other differential core.tab sites for this pair within ±--cluster-window bp.", "Number of pair-specific SNPs in the local window, including this SNP."),
        "pair_clustered_snp": ("pair_local_snp_count compared with --cluster-min-snps.", "yes when this SNP belongs to a local SNP cluster."),
        "removed_by_gubbins": ("gubbins.recombination_predictions.gff coordinates.", "yes when the SNP lies in a predicted recombinant interval; only calculated automatically for a single-contig core.tab."),
        "qc_warnings": ("QC checks on this row's VCF/BAM evidence and context.", "Semicolon-separated flags; they request review and do not remove the row."),
        "qc_status": ("qc_warnings.", "PASS when no QC flags were raised; REVIEW otherwise."),
        "review_reasons": ("qc_warnings.", "QC flags for this SNP, or none."),
    }
    if column in definitions:
        return definitions[column]
    for number in (1, 2):
        prefix = f"sample_{number}_"
        if not column.startswith(prefix):
            continue
        suffix = column[len(prefix):]
        sample_label = f"sample_{number}"
        evidence = {
            "evidence_status": ("VCF/BAM lookup result.", "How evidence was obtained: variant_record, reference_call_mpileup, or a missing-evidence status."),
            "evidence_availability": ("evidence_status.", "available when VCF or BAM evidence was obtained; missing otherwise."),
            "vcf_record_pos": ("snps.vcf: POS.", "Position of the supporting VCF record; blank for BAM-derived reference evidence."),
            "vcf_ref": ("snps.vcf: REF; for BAM reference evidence, core_ref.", "Reference allele of the supporting record."),
            "vcf_alt": ("snps.vcf: ALT; for BAM reference evidence, the other sample's call.", "Alternate allele associated with the comparison."),
            "qual": ("snps.vcf: QUAL.", "Variant quality reported by the caller."),
            "filter": ("snps.vcf: FILTER.", "VCF filter status reported by the caller."),
            "genotype": ("snps.vcf: FORMAT/GT.", "Genotype recorded for this sample."),
            "depth": ("snps.vcf: FORMAT/DP, otherwise INFO/DP; BAM reference evidence: mpileup depth.", "Total read depth used for allele fractions."),
            "ref_count": ("snps.vcf: FORMAT/RO, otherwise INFO/RO; BAM reference evidence: parsed mpileup reference bases.", "Reads supporting the reference allele."),
            "alt_count": ("snps.vcf: FORMAT/AO, otherwise INFO/AO; BAM reference evidence: parsed mpileup bases matching the other sample's call.", "Reads supporting the alternate allele."),
            "variant_type": ("snps.vcf: INFO/TYPE; BAM reference evidence: reference_call.", "Variant type reported by the caller or reference-call label."),
            "ref_forward_count": ("snps.raw.vcf: INFO/SRF; BAM reference evidence: parsed mpileup forward bases.", "Reference-supporting reads on the forward strand."),
            "ref_reverse_count": ("snps.raw.vcf: INFO/SRR; BAM reference evidence: parsed mpileup reverse bases.", "Reference-supporting reads on the reverse strand."),
            "alt_forward_count": ("snps.raw.vcf: INFO/SAF; BAM reference evidence: parsed mpileup forward bases.", "Alternate-supporting reads on the forward strand."),
            "alt_reverse_count": ("snps.raw.vcf: INFO/SAR; BAM reference evidence: parsed mpileup reverse bases.", "Alternate-supporting reads on the reverse strand."),
            "homopolymer_run": ("snps.raw.vcf: INFO/RUN.", "Homopolymer run length reported by FreeBayes."),
            "nearest_indel_distance": ("snps.raw.vcf: positions with INFO/TYPE other than snp.", "Distance in bases to the nearest raw indel or complex call."),
            "ref_fraction": ("ref_count / depth.", "Fraction of depth supporting the reference allele."),
            "alt_fraction": ("alt_count / depth.", "Fraction of depth supporting the alternate allele."),
            "called_allele": (f"core.tab: {sample_label} call.", "The base called for this sample at this SNP."),
            "called_allele_count": ("ref_count when called_allele equals core_ref; otherwise alt_count.", "Read count supporting the called allele."),
            "called_allele_fraction": ("called_allele_count / depth.", "Fraction of depth supporting the allele called in core.tab."),
            "called_allele_forward_count": ("Reference or alternate forward count selected according to called_allele.", "Called-allele reads on the forward strand."),
            "called_allele_reverse_count": ("Reference or alternate reverse count selected according to called_allele.", "Called-allele reads on the reverse strand."),
            "called_allele_minor_strand_fraction": ("If the core.tab call equals core_ref: min(ref_forward_count, ref_reverse_count) / (ref_forward_count + ref_reverse_count); otherwise the same formula using alt_forward_count and alt_reverse_count.", "Fraction of called-allele reads on the less represented strand; lower values indicate stronger strand imbalance."),
            "evidence_summary": ("Called-allele count, depth, fraction and strand counts above.", "Human-readable evidence summary for this sample and SNP."),
        }
        if suffix in evidence:
            return evidence[suffix]
    return ("Column descriptions generated!")


def build_data_dictionary_rows(
    pairs_path: Path, variants_path: Path, summary_path: Path | None
) -> list[dict[str, str]]:
    """Build column descriptions for display in the Excel report only."""
    tables = [(pairs_path.name, pairs_path), (variants_path.name, variants_path)]
    if summary_path is not None:
        tables.append((summary_path.name, summary_path))
    rows = []
    for table_name, table_path in tables:
        with table_path.open(newline="") as table:
            headers = next(csv.reader(table, delimiter="\t"), [])
        for column in headers:
            source, description = column_description(column)
            rows.append(
                {
                    "table": table_name,
                    "column": column,
                    "source_field_or_calculation": source,
                    "meaning": description,
                }
            )
    return rows



def read_tsv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Read a result TSV preserving its header order for a summary Excel worksheet"""
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        return list(reader.fieldnames or []), list(reader)



def excel_cell_value(header: str, value: str) -> int | float | str:
    """
    Convert quantitative result columns to Excel numbers to avoid issues when opening such file
    """
    numeric_columns = {
        "pos", "snp_distance", "comparable_core_sites", "ignored_non_acgt_sites",
        "core_tab_matrix_distance", "phylo_aln_distance", "clean_core_aln_distance",
        "snps_removed_by_gubbins", "comparable_core_fraction", "pair_local_snp_count",
    }
    numeric_suffixes = (
        "_vcf_record_pos", "_qual", "_depth", "_ref_count", "_alt_count",
        "_ref_forward_count", "_ref_reverse_count", "_alt_forward_count",
        "_alt_reverse_count", "_homopolymer_run", "_nearest_indel_distance",
        "_ref_fraction", "_alt_fraction", "_called_allele_count",
        "_called_allele_fraction", "_called_allele_forward_count",
        "_called_allele_reverse_count", "_called_allele_minor_strand_fraction",
    )
    if value in ("", "not_available") or (header not in numeric_columns and not header.endswith(numeric_suffixes)):
        return value
    try:
        numeric_value = float(value)
    except ValueError:
        return value
    return int(numeric_value) if numeric_value.is_integer() else numeric_value



def style_table_worksheet(ws, headers: list[str], rows: list[dict[str, str]]) -> None:
    """Write and format an initial README sheet in the Excel output file, from TSV records"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="4472C4")
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(1, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_number, row in enumerate(rows, start=2):
        for column, header in enumerate(headers, start=1):
            value = excel_cell_value(header, row.get(header, ""))
            cell = ws.cell(row_number, column, value)
            if isinstance(value, (int, float)):
                cell.number_format = "0.0000" if header.endswith("_fraction") else "0"
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if headers:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
        for column, header in enumerate(headers, start=1):
            longest = max([len(header)] + [len(str(row.get(header, ""))) for row in rows])
            ws.column_dimensions[get_column_letter(column)].width = min(max(longest + 2, 12), 45)


def write_excel_report(
    path: Path,
    pairs_path: Path,
    variants_path: Path,
    summary_path: Path | None,
    args: argparse.Namespace,
    sample_count: int,
    core_site_count: int,
) -> None:
    """Create an Excel report with a first-sheet README, using the style defined in style_table_worksheet.
    README explains the role of every sheet, the distance definition, the sources of
    variant evidence, and the most important interpretation caveats.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "Excel report requested but openpyxl is not installed; please install openpyxl "
            "or use --no-excel"
        ) from error

    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    readme.merge_cells("A1:C1")
    readme["A1"] = "Close-pair SNP QC report"
    readme["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    readme["A1"].fill = title_fill
    readme["A1"].alignment = Alignment(vertical="center")
    readme.append([])
    readme.append(["Sheet", "What it contains", "How to use it"])
    for cell in readme[3]:
        cell.font = Font(bold=True)
        cell.fill = section_fill
    sheet_entries = [
        ("README", "This index of the workbook sheets.", "Start here."),
        ("Close pairs", "One row per pair of samples retained below the requested SNP threshold.", "Use it to identify close pairs and inspect distance-matrix consistency."),
        ("SNP variants", "One technical row for each SNP that differs within a retained pair.", "Use it for detailed VCF, BAM and QC evidence."),
    ]
    if summary_path is not None:
        sheet_entries.append(("Variant summary", "A compact version of the SNP-variant table.", "Use it for manual evaluation; consult SNP variants sheet to check all fields."))
    sheet_entries.append(("Column descriptions", "Definition and exact source field or calculation for every output column.", "Check this sheet to find the meaning of each column field."))
    for entry in sheet_entries:
        readme.append(entry)
    for row in readme.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column, width in enumerate((28, 58, 65), start=1):
        readme.column_dimensions[get_column_letter(column)].width = width
    readme.freeze_panes = "A4"

    sheets = [
        ("Close pairs", pairs_path),
        ("SNP variants", variants_path),
    ]
    if summary_path is not None:
        sheets.append(("Variant summary", summary_path))
    for sheet_name, table_path in sheets:
        worksheet = workbook.create_sheet(sheet_name)
        headers, rows = read_tsv_rows(table_path)
        style_table_worksheet(worksheet, headers, rows)
    dictionary = workbook.create_sheet("Column descriptions")
    style_table_worksheet(
        dictionary,
        ["table", "column", "source_field_or_calculation", "meaning"],
        build_data_dictionary_rows(pairs_path, variants_path, summary_path),
    )
    workbook.save(path)



def sha256_file(path: Path) -> str | None:
    """Return a content hash for a present input file, otherwise ``None``."""
    if not path.is_file():
        return None
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def write_metadata(path: Path, args: argparse.Namespace, core_tab: Path, pairs_path: Path, variants_path: Path, summary_path: Path | None, excel_path: Path | None) -> None:
    """Write provenance sufficient to interpret and reproduce a result table."""
    metadata = {
        "script": "evaluate_close_pairs.py",
        "created_utc": datetime.now(timezone.utc).isoformat(),
        "python_version": sys.version,
        "parameters": {
            key: str(value.resolve()) if isinstance(value, Path) else value
            for key, value in vars(args).items()
        },
        "inputs": {"core_tab": {"path": str(core_tab), "sha256": sha256_file(core_tab)}},
        "outputs": [str(pairs_path), str(variants_path)] + ([str(summary_path)] if summary_path else []) + ([str(excel_path)] if excel_path else []),
    }
    path.write_text(json.dumps(metadata, indent=2) + "\n")


def main() -> int:
    """Run close-pair detection, evidence collection, QC and report generation."""
    args = parse_args()
    if args.threshold <= 0:
        raise ValueError("--threshold must be greater than zero")
    if args.pileup_mapq < 0 or args.pileup_baseq < 0:
        raise ValueError("pileup quality thresholds cannot be negative")
    if args.warn_min_depth < 0:
        raise ValueError("--warn-min-depth cannot be negative")
    if not 0 <= args.warn_min_called_fraction <= 1:
        raise ValueError("--warn-min-called-fraction must be between 0 and 1")
    if not 0 <= args.warn_min_strand_fraction <= 0.5:
        raise ValueError("--warn-min-strand-fraction must be between 0 and 0.5")
    if args.warn_indel_distance < 0 or args.warn_homopolymer_run < 1:
        raise ValueError(
            "indel distance must be non-negative and homopolymer run positive"
        )
    if args.cluster_window < 1 or args.cluster_min_snps < 2:
        raise ValueError(
            "cluster window must be positive and --cluster-min-snps at least 2"
        )
    snippy_dir = args.snippy_dir.resolve()
    core_tab = (args.core_tab or snippy_dir / "core.tab").resolve()
    samples, sites = read_core_tab(core_tab)
    pairs = find_close_pairs(samples, sites, args.threshold)

    chromosomes = {chrom for chrom, *_ in sites}
    gubbins_path = (
        args.gubbins_gff or snippy_dir / "gubbins.recombination_predictions.gff"
    ).resolve()
    gubbins_intervals = []
    if gubbins_path.is_file():
        if len(chromosomes) == 1:
            gubbins_intervals = read_gubbins_intervals(gubbins_path)
        else:
            print(
                "WARNING: Gubbins coordinates cannot be mapped automatically to a "
                "multi-contig core.tab; removed variant positions will be unavailable",
                file=sys.stderr,
            )
    else:
        print(f"WARNING: Gubbins GFF not found: {gubbins_path}", file=sys.stderr)

    matrices_dir = (args.matrices_dir or snippy_dir).resolve()
    matrix_specs = (
        ("core_tab_matrix_distance", "core_tab_snp_distances.tsv"),
        ("phylo_aln_distance", "phylo_snp_distances.tsv"),
        ("clean_core_aln_distance", "clean_core_snp_distances.tsv"),
    )
    matrices = {}
    for column, filename in matrix_specs:
        matrix_path = matrices_dir / filename
        if matrix_path.is_file():
            matrices[column] = read_distance_matrix(matrix_path)
        else:
            print(
                f"WARNING: comparison matrix not found: {matrix_path}", file=sys.stderr
            )
            matrices[column] = {}
    for pair in pairs:
        key = frozenset((pair["sample_1"], pair["sample_2"]))
        for column, _ in matrix_specs:
            pair[column] = matrices[column].get(key, "")
        phylo_distance = pair["phylo_aln_distance"]
        clean_distance = pair["clean_core_aln_distance"]
        pair["core_tab_matrix_matches"] = (
            "yes"
            if pair["core_tab_matrix_distance"] == pair["snp_distance"]
            else ("" if pair["core_tab_matrix_distance"] == "" else "no")
        )
        pair["phylo_matches_core_tab"] = (
            "yes"
            if phylo_distance == pair["snp_distance"]
            else ("" if phylo_distance == "" else "no")
        )
        pair["snps_removed_by_gubbins"] = (
            pair["snp_distance"] - clean_distance if clean_distance != "" else ""
        )
        pair["clean_distance_valid"] = (
            "yes"
            if clean_distance != "" and clean_distance <= pair["snp_distance"]
            else ("" if clean_distance == "" else "no")
        )
        comparison_warnings = []
        if pair["core_tab_matrix_matches"] == "no":
            comparison_warnings.append("core_tab_matrix_mismatch")
        if pair["phylo_matches_core_tab"] == "no":
            comparison_warnings.append("phylo_core_tab_mismatch")
        if pair["clean_distance_valid"] == "no":
            comparison_warnings.append("clean_distance_greater_than_core")
        if any(pair[column] == "" for column, _ in matrix_specs):
            comparison_warnings.append("comparison_matrix_missing")
        pair["distance_comparison_warnings"] = ";".join(comparison_warnings)

    needed_samples = sorted({p[key] for p in pairs for key in ("sample_1", "sample_2")})
    vcf_indexes = {}
    non_snp_positions = {}
    for sample in needed_samples:
        path = locate_vcf(snippy_dir, sample)
        if path is None:
            message = f"No snps.vcf or snps.vcf.gz found for sample {sample}"
            if args.missing_vcf == "error":
                raise FileNotFoundError(message)
            print(f"WARNING: {message}", file=sys.stderr)
            vcf_indexes[sample] = {}
            non_snp_positions[sample] = {}
        else:
            vcf_indexes[sample] = read_vcf(path)
            raw_path = locate_raw_vcf(snippy_dir, sample)
            if raw_path is not None:
                raw_index = read_vcf(raw_path)
                contextual_fields = (
                    "ref_forward_count",
                    "ref_reverse_count",
                    "alt_forward_count",
                    "alt_reverse_count",
                    "homopolymer_run",
                )
                for key, evidence in vcf_indexes[sample].items():
                    raw_evidence = raw_index.get(key, {})
                    for field in contextual_fields:
                        evidence[field] = raw_evidence.get(field, "")
                non_snp_positions[sample] = read_non_snp_positions(raw_path)
            else:
                print(
                    f"WARNING: no snps.raw.vcf(.gz) for {sample}; strand, "
                    "homopolymer, and nearby-indel checks will be unavailable",
                    file=sys.stderr,
                )
                non_snp_positions[sample] = {}

    for sample, index in vcf_indexes.items():
        sample_non_snps = non_snp_positions[sample]
        for (chrom, pos, _), evidence in index.items():
            distance = nearest_distance(sample_non_snps.get(chrom, []), pos)
            evidence["nearest_indel_distance"] = "" if distance is None else distance

    bam_indexes = {sample: {} for sample in needed_samples}
    if not args.no_bam_evidence:
        samtools = (
            shutil.which(args.samtools)
            if Path(args.samtools).name == args.samtools
            else args.samtools
        )
        if not samtools:
            raise FileNotFoundError(
                f"samtools executable not found: {args.samtools!r}; load samtools, use "
                "--samtools /path/to/samtools, or use --no-bam-evidence"
            )
        reference_positions = {sample: set() for sample in needed_samples}
        for pair in pairs:
            for chrom, pos, ref, call_1, call_2 in pair["differences"]:
                if call_1 == ref:
                    reference_positions[pair["sample_1"]].add((chrom, pos))
                if call_2 == ref:
                    reference_positions[pair["sample_2"]].add((chrom, pos))
        for sample, positions in reference_positions.items():
            if not positions:
                continue
            bam = snippy_dir / sample / "snps.bam"
            if not bam.is_file():
                raise FileNotFoundError(f"BAM not found for sample {sample}: {bam}")
            reference = (
                args.reference.resolve()
                if args.reference
                else snippy_dir / sample / "reference" / "ref.fa"
            )
            if not reference.is_file():
                raise FileNotFoundError(f"Reference FASTA not found: {reference}")
            bam_indexes[sample] = read_bam_pileup(
                str(samtools),
                reference,
                bam,
                positions,
                args.pileup_mapq,
                args.pileup_baseq,
            )
            for (chrom, pos), pileup in bam_indexes[sample].items():
                distance = nearest_distance(
                    non_snp_positions[sample].get(chrom, []), pos
                )
                pileup["nearest_indel_distance"] = "" if distance is None else distance

    prefix = args.output_prefix
    if not prefix.is_absolute():
        prefix = Path.cwd() / prefix
    prefix.parent.mkdir(parents=True, exist_ok=True)
    pairs_path = prefix.with_name(prefix.name + "_pairs.tsv")
    variants_path = prefix.with_name(prefix.name + "_variants.tsv")

    pair_columns = [
        "pair_id",
        "sample_1",
        "sample_2",
        "snp_distance",
        "comparable_core_sites",
        "ignored_non_acgt_sites",
        "core_tab_matrix_distance",
        "core_tab_matrix_matches",
        "phylo_aln_distance",
        "phylo_matches_core_tab",
        "clean_core_aln_distance",
        "snps_removed_by_gubbins",
        "clean_distance_valid",
        "distance_comparison_warnings",
    ]
    with pairs_path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=pair_columns, delimiter="\t")
        writer.writeheader()
        for number, pair in enumerate(pairs, 1):
            writer.writerow(
                {
                    "pair_id": f"pair_{number:04d}",
                    **{k: pair[k] for k in pair_columns[1:]},
                }
            )

    base_columns = [
        "pair_id",
        "sample_1",
        "sample_2",
        "snp_distance",
        "chrom",
        "pos",
        "core_ref",
        "sample_1_call",
        "sample_2_call",
        "pair_local_snp_count",
        "pair_clustered_snp",
        "removed_by_gubbins",
        "qc_warnings",
    ]
    evidence_output = []
    for n in (1, 2):
        evidence_output.extend(
            [f"sample_{n}_evidence_status"]
            + [f"sample_{n}_{key}" for key in EVIDENCE_COLUMNS]
            + [f"sample_{n}_ref_fraction", f"sample_{n}_alt_fraction"]
        )
    qc_warnings: dict[tuple[str, str, int, str], tuple[str, str]] = {}
    with variants_path.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=base_columns + evidence_output, delimiter="\t"
        )
        writer.writeheader()
        for number, pair in enumerate(pairs, 1):
            for chrom, pos, ref, call_1, call_2 in pair["differences"]:
                local_snp_count = sum(
                    other_chrom == chrom and abs(other_pos - pos) <= args.cluster_window
                    for other_chrom, other_pos, *_ in pair["differences"]
                )
                row = {
                    "pair_id": f"pair_{number:04d}",
                    "sample_1": pair["sample_1"],
                    "sample_2": pair["sample_2"],
                    "snp_distance": pair["snp_distance"],
                    "chrom": chrom,
                    "pos": pos,
                    "core_ref": ref,
                    "sample_1_call": call_1,
                    "sample_2_call": call_2,
                    "pair_local_snp_count": local_snp_count,
                    "pair_clustered_snp": (
                        "yes" if local_snp_count >= args.cluster_min_snps else "no"
                    ),
                    "removed_by_gubbins": (
                        "yes" if position_in_intervals(pos, gubbins_intervals) else "no"
                    ),
                }
                row_warning_types = []
                if row["removed_by_gubbins"] == "yes":
                    row_warning_types.append("gubbins_removed")
                if local_snp_count >= args.cluster_min_snps:
                    row_warning_types.append("local_snp_cluster")
                    message = (
                        f"pair={row['pair_id']} samples={pair['sample_1']},{pair['sample_2']} "
                        f"position={chrom}:{pos} local_snps={local_snp_count} "
                        f"window=+/-{args.cluster_window}"
                    )
                    key = ("local_snp_cluster", str(row["pair_id"]), pos, chrom)
                    qc_warnings.setdefault(key, ("local_snp_cluster", message))
                for n, call, other_call in ((1, call_1, call_2), (2, call_2, call_1)):
                    sample = pair[f"sample_{n}"]
                    status, evidence = evidence_for(
                        vcf_indexes[sample],
                        bam_indexes[sample],
                        chrom,
                        pos,
                        ref,
                        call,
                        other_call,
                    )
                    row[f"sample_{n}_evidence_status"] = status
                    row.update(
                        {f"sample_{n}_{key}": value for key, value in evidence.items()}
                    )
                    add_fractions(row, f"sample_{n}")
                    sample_warnings = collect_qc_warnings(
                        row,
                        n,
                        args.warn_min_depth,
                        args.warn_min_called_fraction,
                        args.warn_min_strand_fraction,
                        args.warn_indel_distance,
                        args.warn_homopolymer_run,
                    )
                    for warning_type, message in sample_warnings:
                        row_warning_types.append(f"sample_{n}:{warning_type}")
                        # A sample/site can occur in several close pairs; report it once.
                        key = (warning_type, sample, pos, chrom)
                        qc_warnings.setdefault(key, (warning_type, message))
                # Preserve warning order while removing repeated labels.
                row["qc_warnings"] = ";".join(dict.fromkeys(row_warning_types))
                writer.writerow(row)

    variant_rows = enhance_output_tables(pairs_path, variants_path, args.threshold)
    summary_path = None
    if not args.no_summary:
        summary_path = prefix.with_name(prefix.name + "_variants_summary.tsv")
        write_variant_summary(summary_path, variant_rows)
    excel_path = None
    if not args.no_excel:
        excel_path = args.excel_output or prefix.with_name(prefix.name + "_report.xlsx")
        if not excel_path.is_absolute():
            excel_path = Path.cwd() / excel_path
        try:
            write_excel_report(
                excel_path, pairs_path, variants_path, summary_path,
                args, len(samples), len(sites),
            )
        except RuntimeError as error:
            print(f"WARNING: {error}", file=sys.stderr)
            excel_path = None
    metadata_path = None
    if not args.no_metadata:
        metadata_path = prefix.with_name(prefix.name + "_metadata.json")
        write_metadata(metadata_path, args, core_tab, pairs_path, variants_path, summary_path, excel_path)

    variant_count = sum(len(pair["differences"]) for pair in pairs)
    print(f"Samples: {len(samples)}")
    print(f"Core SNP sites: {len(sites)}")
    print(f"Pairs with < {args.threshold} SNPs: {len(pairs)}")
    print(f"Pair-specific variant rows: {variant_count}")
    print(f"Wrote: {pairs_path}")
    print(f"Wrote: {variants_path}")
    if summary_path:
        print(f"Wrote: {summary_path}")
    if excel_path:
        print(f"Wrote: {excel_path}")
    if metadata_path:
        print(f"Wrote: {metadata_path}")
    if not args.no_qc_warnings:
        for warning_type, message in qc_warnings.values():
            print(f"WARNING [{warning_type}] {message}")
    warning_counts = {}
    for warning_type, _ in qc_warnings.values():
        warning_counts[warning_type] = warning_counts.get(warning_type, 0) + 1
    if warning_counts:
        summary = ", ".join(
            f"{key}={value}" for key, value in sorted(warning_counts.items())
        )
        print(f"Unique QC warnings: {summary}")
    else:
        print("Unique QC warnings: none")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (OSError, RuntimeError, ValueError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(2)
